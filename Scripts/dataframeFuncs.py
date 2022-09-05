import os
import shutil
from functools import reduce

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def ExcelToDataframe(filename: str)-> pd.DataFrame:
    '''convert excel file to dataframe'''
    return pd.read_excel(filename)

def CSVtoDataframe(filename: str)-> pd.DataFrame:
    '''convert csv tile to dataframe'''
    return pd.read_csv(filename)

def deleteTitleRows(df: pd.DataFrame)-> pd.DataFrame:
    """remove rows that correspond to the column titles for each group"""

    df = df[df["Name"].str.contains("Name") == False]
    return df

def deleteBlankRows(df: pd.DataFrame)-> pd.DataFrame:
    """remove blank rows"""

    df.dropna(subset=["Name"],inplace=True)
    return df

def AppendTwoFrames(frame1: pd.DataFrame, frame2: pd.DataFrame)-> pd.DataFrame:
    """append two frames together"""

    return pd.concat([frame1, frame2], sort=False)

def CombineFrames(folder: str)-> pd.DataFrame:
    """Combine the workbooks in the folder to one frame"""
    print("Combining boards")

    for dirpath, dirnames, files in os.walk(folder):
        #creating a list of the dataframes for each board
        frames = [pd.read_excel(dirpath + filename) for filename in files]
        #reducing all of the frames into one frame by concat
        frame = reduce(AppendTwoFrames, frames)
        #removing the extra title rows
        frame = frame[frame["Name"].str.contains("Name") == False]

        return frame

def FrametoWorkbook(frame: pd.DataFrame)-> Workbook:
    """convert frame to workbook"""
    board = Workbook()
    boardsheet = board.active
    for row in dataframe_to_rows(frame, index=False, header=True):
        boardsheet.append(row)
    return board

def SaveFrame(frame: pd.DataFrame, title: str, folderPath: str, fileID: str = ".xlsx"):
    """Save Frame to excel sheet"""

    wb = FrametoWorkbook(frame)

    filename = title + fileID

    wb.save(filename=filename)

    print(title + " saved in file: " + filename)

    shutil.move(filename, folderPath)
